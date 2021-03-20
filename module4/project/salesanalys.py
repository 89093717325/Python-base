import pandas
import collections
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from itertools import chain

ALPHABET = 'ABCDEFGHIJKLMNOPQRSTUFV'
browsers_list = [] 
items_list = [] 

excel_logs = pandas.read_excel('logs.xlsx', sheet_name='log')
excel_logs_dict = excel_logs.to_dict(orient='records')

#Создадим списки посещений браузеров и покупок 
for log in excel_logs_dict:
    browsers_list.append(log['Браузер'])
    items_list.append(log['Купленные товары'].split(','))

items_list_final = list(chain.from_iterable(items_list))

#Удаляем из списка элементы: Ещё 2 варианта и Ещё 3 варианта
for i in items_list_final:
    if i == 'Ещё 2 варианта':
        items_list_final.remove(i)

for i in items_list_final:
    if i == 'Ещё 3 варианта':
        items_list_final.remove(i)

#Выводим 7 самых популярных браузеров и товаров
browsers_counter = collections.Counter(browsers_list)
items_counter = collections.Counter(items_list_final)
top_browsers = browsers_counter.most_common(7)
top_items = items_counter.most_common(7)
print(f'Список топ 7 браузеров: {top_browsers}')
print(f'Список топ 7 товаров: {top_items}')

#Создадим списки браузеров и покупок по месяцам
browsers_by_month = defaultdict(int)
items_by_month = defaultdict(int)
for log in excel_logs_dict:
    for j in log['Купленные товары'].split(','):
        items_by_month[j.strip() + str(log['Дата посещения'].month)] += 1
    browsers_by_month[log['Браузер'] + str(log['Дата посещения'].month)] += 1
print(items_by_month)
print(browsers_by_month)

#Вычисляем самые популярные и самые не популярные товары среди м и ж
male_items = []
for log in excel_logs_dict:
    if log['Пол'] == 'м':
        for j in log['Купленные товары'].split(','):
            male_items.append(j)
male_items_counter = collections.Counter(male_items)
male_items_top = male_items_counter.most_common(1)
male_items_low = male_items_counter.most_common()[:-(len(male_items_counter)+1):-1][1]
print(f'Самый популярный товар среди мужчин: {male_items_top}')
print(f'Самый не популярный товар среди мужчин: {male_items_low}')

female_items = []
for log in excel_logs_dict:
    if log['Пол'] == 'ж':
        for j in log['Купленные товары'].split(','):
            female_items.append(j)
female_items_counter = collections.Counter(female_items)
female_items_top = female_items_counter.most_common(1)
female_items_low = female_items_counter.most_common()[:-(len(female_items_counter)+1):-1][1]
print(f'Самый популярный товар среди женщин: {female_items_top}')
print(f'Самый не популярный товар среди женщин: {female_items_low}')

"""
#Записываем полученные данные в файл "report"
wb = load_workbook(filename='report.xlsx')
sheet = wb['Sheet1']


for i in range(5, 12):
    sheet.cell(row = i, column = 1).value = top_browsers[i - 5][0]

for i in range(19, 26):
    sheet.cell(row = i, column = 1).value = top_items[i - 19][0]

row = 5
for item in top_browsers:
    col = 2
    for i in range(1, 13):
        sheet[ALPHABET[col]+str(row)] = browsers_by_month.get(item[0]+str(i), '')
        col += 1
    row += 1

row = 19
for item in top_items:
    col = 2
    for i in range(1, 13):
        sheet[ALPHABET[col]+str(row)] = items_by_month.get(item[0]+str(i), 0)
        col += 1
    row += 1

sheet['B31'] = male_items_top[0][0]
sheet['B32'] = female_items_top[0][0]
sheet['B33'] = male_items_low[0]
sheet['B34'] = female_items_low[0]

wb.save('report.xlsx')"""