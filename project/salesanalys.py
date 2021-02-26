import pandas
import collections
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from itertools import chain

excel_logs = pandas.read_excel('logs.xlsx', sheet_name='log')
excel_logs_dict = excel_logs.to_dict(orient='records')#Преобразуем данные в список

#Создадим списки посещений браузеров и покупок 
browsers_list = [] 
items_list = [] 
for log in excel_logs_dict:
    browsers_list.append(log['Браузер'])
    items_list.append(log['Купленные товары'].split(','))

items_list_final = list(chain.from_iterable(items_list))

#Удаляем из списка элементы: Ещё 2 варианта и Ещё 3 варианта
for i in items_list_final:
    if i == 'Ещё 2 варианта':
        items_list_final.remove(i)

for i in items_list_final:
    if i == 'Ещё 3 варианта':
        items_list_final.remove(i)

#В объект Counter передаем список браузеров 
#Для товаров тоже самое
browsers_counter = collections.Counter(browsers_list)
items_counter = collections.Counter(items_list_final)


#Выводим 7 самых популярных браузеров и товаров
top_browsers = browsers_counter.most_common(7)
top_items = items_counter.most_common(7)
print(f'Список топ 7 браузеров: {top_browsers}')
print(f'Список топ 7 товаров: {top_items}')

#здесь начинается первая часть моего колхозного кода
date_list1 = []
for log in excel_logs_dict:
    if log['Браузер'] == top_browsers[0][0]:
        date_list1.append(log['Дата посещения'].month)
date1_counter = collections.Counter(date_list1)
print(f'Количество посещения браузера №1 по месяцам: {date1_counter}')

date_list2 = []
for log in excel_logs_dict:
    if log['Браузер'] == top_browsers[1][0]:
        date_list2.append(log['Дата посещения'].month)
date2_counter = collections.Counter(date_list2)
print(f'Количество посещения браузера №2 по месяцам: {date2_counter}')

date_list3 = []
for log in excel_logs_dict:
    if log['Браузер'] == top_browsers[2][0]:
        date_list3.append(log['Дата посещения'].month)
date3_counter = collections.Counter(date_list3)
print(f'Количество посещения браузера №3 по месяцам: {date3_counter}')

date_list4 = []
for log in excel_logs_dict:
    if log['Браузер'] == top_browsers[3][0]:
        date_list4.append(log['Дата посещения'].month)
date4_counter = collections.Counter(date_list4)
print(f'Количество посещения браузера №4 по месяцам: {date4_counter}')

date_list5 = []
for log in excel_logs_dict:
    if log['Браузер'] == top_browsers[4][0]:
        date_list5.append(log['Дата посещения'].month)
date5_counter = collections.Counter(date_list5)
print(f'Количество посещения браузера №5 по месяцам: {date5_counter}')

date_list6 = []
for log in excel_logs_dict:
    if log['Браузер'] == top_browsers[5][0]:
        date_list6.append(log['Дата посещения'].month)
date6_counter = collections.Counter(date_list6)
print(f'Количество посещения браузера №6 по месяцам: {date6_counter}')

date_list7 = []
for log in excel_logs_dict:
    if log['Браузер'] == top_browsers[6][0]:
        date_list7.append(log['Дата посещения'].month)
date7_counter = collections.Counter(date_list7)
print(f'Количество посещения браузера №7 по месяцам: {date7_counter}')

#Здесь начинается вторая часть моего колхозного кода
date_item_list1 = []
for log in excel_logs_dict:
    for j in log['Купленные товары'].split(','):
        if j == top_items[0][0]:
            date_item_list1.append(log['Дата посещения'].month)
date_item1_counter = collections.Counter(date_item_list1)
print(f'Количество покупок товара №1 по месяцам: {date_item1_counter}')

date_item_list2 = []
for log in excel_logs_dict:
    for j in log['Купленные товары'].split(','):
        if j == top_items[1][0]:
            date_item_list2.append(log['Дата посещения'].month)
date_item2_counter = collections.Counter(date_item_list2)
print(f'Количество покупок товара №2 по месяцам: {date_item2_counter}')

date_item_list3 = []
for log in excel_logs_dict:
    for j in log['Купленные товары'].split(','):
        if j == top_items[2][0]:
            date_item_list3.append(log['Дата посещения'].month)
date_item3_counter = collections.Counter(date_item_list3)
print(f'Количество покупок товара №3 по месяцам: {date_item3_counter}')

date_item_list4 = []
for log in excel_logs_dict:
    for j in log['Купленные товары'].split(','):
        if j == top_items[3][0]:
            date_item_list4.append(log['Дата посещения'].month)
date_item4_counter = collections.Counter(date_item_list4)
print(f'Количество покупок товара №4 по месяцам: {date_item4_counter}')

date_item_list5 = []
for log in excel_logs_dict:
    for j in log['Купленные товары'].split(','):
        if j == top_items[4][0]:
            date_item_list5.append(log['Дата посещения'].month)
date_item5_counter = collections.Counter(date_item_list5)
print(f'Количество покупок товара №5 по месяцам: {date_item5_counter}')

date_item_list6 = []
for log in excel_logs_dict:
    for j in log['Купленные товары'].split(','):
        if j == top_items[5][0]:
            date_item_list6.append(log['Дата посещения'].month)
date_item6_counter = collections.Counter(date_item_list6)
print(f'Количество покупок товара №6 по месяцам: {date_item6_counter}')

date_item_list7 = []
for log in excel_logs_dict:
    for j in log['Купленные товары'].split(','):
        if j == top_items[6][0]:
            date_item_list7.append(log['Дата посещения'].month)
date_item7_counter = collections.Counter(date_item_list7)
print(f'Количество покупок товара №7 по месяцам: {date_item7_counter}')

#Вычисляем самые популярные и самые не популярные товары среди м и ж
male_items = []
for log in excel_logs_dict:
    if log['Пол'] == 'м':
        for j in log['Купленные товары'].split(','):
            male_items.append(j)
male_items_counter = collections.Counter(male_items)
male_items_top = male_items_counter.most_common(1)
male_items_low = male_items_counter.most_common()[:-(len(male_items_counter)+1):-1][1]
print(f'Самый популярный товар среди мужчин: {male_items_top}')
print(f'Самый не популярный товар среди мужчин: {male_items_low}')

female_items = []
for log in excel_logs_dict:
    if log['Пол'] == 'ж':
        for j in log['Купленные товары'].split(','):
            female_items.append(j)
female_items_counter = collections.Counter(female_items)
female_items_top = female_items_counter.most_common(1)
female_items_low = female_items_counter.most_common()[:-(len(female_items_counter)+1):-1][1]
print(f'Самый популярный товар среди женщин: {female_items_top}')
print(f'Самый не популярный товар среди женщин: {female_items_low}')

#Записываем полученные данные в файл "report"
wb = load_workbook(filename='report.xlsx')
sheet = wb['Sheet1']

sheet['B31'] = male_items_top[0][0]
sheet['B32'] = female_items_top[0][0]
sheet['B33'] = male_items_low[0]
sheet['B34'] = female_items_low[0]

for i in range(5, 12):
    sheet.cell(row = i, column = 1).value = top_browsers[i - 5][0]

for i in range(19, 26):
    sheet.cell(row = i, column = 1).value = top_items[i - 19][0]

for i in range(3, 15):
    sheet.cell(row = 5, column = i).value = date1_counter[i - 2]
    sheet.cell(row = 6, column = i).value = date2_counter[i - 2]
    sheet.cell(row = 7, column = i).value = date3_counter[i - 2]
    sheet.cell(row = 8, column = i).value = date4_counter[i - 2]
    sheet.cell(row = 9, column = i).value = date5_counter[i - 2]
    sheet.cell(row = 10, column = i).value = date6_counter[i - 2]
    sheet.cell(row = 11, column = i).value = date7_counter[i - 2]
    sheet.cell(row = 19, column = i).value = date_item1_counter[i - 2]
    sheet.cell(row = 20, column = i).value = date_item2_counter[i - 2]
    sheet.cell(row = 21, column = i).value = date_item3_counter[i - 2]
    sheet.cell(row = 22, column = i).value = date_item4_counter[i - 2]
    sheet.cell(row = 23, column = i).value = date_item5_counter[i - 2]
    sheet.cell(row = 24, column = i).value = date_item6_counter[i - 2]
    sheet.cell(row = 25, column = i).value = date_item7_counter[i - 2]

wb.save('report.xlsx')
